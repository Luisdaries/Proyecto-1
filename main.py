import os
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook


def main():
    print("inicio")

    espacio()

    path = ruta()
    data = Analizar_Folder(path)

    # print(data)
    limpiar_terminal()

    # print(str(path)+"/"+"".join(data))

    # leer_archivos(path, data)
    j = 2
    leer_productos(path,data,j)


def limpiar_terminal():
    os.system("cls")


def ruta():
    print("Ventana de dialogo para la obtencion de la ruta")
    ruta = filedialog.askdirectory()
    return ruta


def Analizar_Folder(ruta):

    data = ruta

    contenido = os.listdir(data)

    Excel = []
    for fichero in contenido:
        if os.path.isfile(os.path.join(data, fichero)) and fichero.endswith('.xlsm'):
            Excel.append(fichero)
    return (Excel)


def espacio():
    print(" ")


def leer_archivos(ruta, archivos):
    # Funcion para obtener los primeros datos de informacion

    # crear libro de excel y descargar

    base = Workbook()
    sheet = base.active

    sheet.title = "Base de datos completa"

    # importar archivo y asignar hoja
    for archivo in archivos:
        wb = openpyxl.load_workbook(str(ruta)+"/"+str(archivo), data_only=True)
        print(archivos.index(archivo))
        print(archivo)


        if hoja != wb["INFO"]:
            continue
        else:
            hoja = wb["INFO"]

        # Crea lista y almacenar valores
        lista = []

        for x in range(10, 17):
            for y in range(2, 5):

                valor = hoja.cell(row=x, column=y).value
                lista.append(valor)

        # Eliminar basura (None)
        lista_limpia = []
        for x in lista:
            if x != None:
                lista_limpia.append(x)

        i = 1
        # Acomodar lista
        for x in range(len(lista_limpia)):
            
            if x % 2 == 0:
                sheet.cell(1, i, lista_limpia[x])
                sheet.cell(archivos.index(archivo)+2, i, lista_limpia[x+1])
                i = i + 1
            sheet.cell(archivos.index(archivo)+2, i, archivo)
            
            base.save("resultado/"+"data.xlsx")

        
def leer_productos(ruta, archivos,j):
# Funcion para obtener los productos

    # crear libro de excel y descargar

    productos = Workbook()
    sheet = productos.active

    sheet.title = "Base de datos completa"

    # importar archivo y asignar hoja
    for archivo in archivos:
        wb = openpyxl.load_workbook(str(ruta)+"/"+str(archivo),data_only=True)
        print(archivos.index(archivo))
        print(archivo)
        hoja = wb["INFO"]

        i = 1          

        limite = 0
        fila = 0
        for col in hoja.iter_rows(min_col=2, min_row=10,  max_col=2, max_row=16):
            for cell in col:
                if cell.value != None:
                    limite = limite + 1
        
        for row in hoja.iter_cols(min_row=10, max_row=10):
            
            for cell in row:
                if cell.value != None:
                    # sheet.cell(i, fila+1, cell.value)
                    fila = fila + 1

        # for col in hoja.iter_rows(min_col=2, min_row=10,  max_col=fila+1, max_row=limite+20,values_only = False):
        for col in hoja.iter_rows(min_col=2, min_row=10,  max_col=4, max_row=16,values_only = False):
            
            for cell in col:
                    if cell.value != "#VALUE!":
                            if cell.value == None:
                                sheet.cell(j, i, "sin dato")
                            else:    
                                sheet.cell(j, i, cell.value)

                    i = i + 1
                    sheet.cell(j, i, archivo)
            j = j + 1
            productos.save("resultado/"+"productos3.xlsx")

            i = 1

main()
